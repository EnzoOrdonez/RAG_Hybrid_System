"""TEMP helper: export professor's 4 tables (xlsx 4 sheets + md) from
exp10_retrieval194 retrieval_metrics.json. Not committed."""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "experiments" / "results" / "exp10_retrieval194"
OUT_DIR = ROOT / "output" / "tables" / "tablas_profe_nota2"
OUT_DIR.mkdir(parents=True, exist_ok=True)

metrics = json.loads((SRC / "retrieval_metrics.json").read_text(encoding="utf-8"))
run = json.loads((SRC / "results.json").read_text(encoding="utf-8"))
sysd = metrics["systems"]

# metric key -> column header
COLS = [
    ("precision@1_mean", "Precision@1"),
    ("precision@5_mean", "Precision@5"),
    ("recall@5_mean", "Recall@5"),
    ("mrr_mean", "MRR"),
    ("ndcg@5_mean", "NDCG@5"),
]
MODELS = ["Mistral 7B", "Qwen 2.5 7B", "Llama 3.1 8B"]

# scenario sheet name -> source system key in retrieval_metrics.json (None = Sin RAG)
SCENARIOS = [
    ("Sin RAG", None),
    ("RAG léxico (BM25)", "RAG Lexico (BM25)"),
    ("RAG denso (BGE)", "RAG Semantico (Dense)"),
    ("RAG híbrido (propuesto)", "RAG Hibrido Propuesto"),
]

NOTA = (
    "Nota. Las métricas de recuperación son independientes del modelo generador, "
    "pues la etapa de recuperación ocurre antes de la generación; por ello los "
    "valores coinciden entre modelos. El escenario sin RAG no recupera documentos, "
    "por lo que estas métricas no aplican. Resultados sobre el corpus reconstruido "
    "(2 697 documentos / 24 481 fragmentos) con las 194 consultas del conjunto de "
    "evaluación depurado."
)


def fmt(x):
    """3 decimals, comma decimal separator -> '0,820'."""
    return f"{x:.3f}".replace(".", ",")


def row_values(system_key):
    """Return the 5 formatted metric strings for a scenario (or '—' x5)."""
    if system_key is None:
        return ["—"] * len(COLS)
    s = sysd[system_key]
    return [fmt(s[k]) for k, _ in COLS]


# ---------------- XLSX ----------------
wb = Workbook()
wb.remove(wb.active)

hdr_fill = PatternFill("solid", fgColor="2F5496")
hdr_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for sheet_name, system_key in SCENARIOS:
    ws = wb.create_sheet(title=sheet_name[:31])
    headers = ["Modelo"] + [h for _, h in COLS]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border

    vals = row_values(system_key)
    for r, model in enumerate(MODELS, start=2):
        ws.cell(row=r, column=1, value=model).alignment = Alignment(
            horizontal="left", vertical="center")
        ws.cell(row=r, column=1).border = border
        for c, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = center
            cell.border = border

    # column widths
    ws.column_dimensions["A"].width = 16
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 13

    # footnote: blank row then merged note across all columns
    note_row = 2 + len(MODELS) + 1
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=len(headers))
    nc = ws.cell(row=note_row, column=1, value=NOTA)
    nc.alignment = left
    nc.font = Font(italic=True, size=9)
    ws.row_dimensions[note_row].height = 64

xlsx_path = OUT_DIR / "tablas_profe_nota2.xlsx"
wb.save(xlsx_path)

# ---------------- MD ----------------
lines = []
lines.append("# Métricas de recuperación — corpus reconstruido (194 consultas)\n")
lines.append(
    f"_Fuente: `experiments/results/exp10_retrieval194/` — "
    f"{run['num_queries']} consultas, corpus {run['corpus']['docs']} docs / "
    f"{run['corpus']['chunks']} fragmentos, seed={run['seed']}._\n"
)
for sheet_name, system_key in SCENARIOS:
    lines.append(f"\n## {sheet_name}\n")
    header = "| Modelo | " + " | ".join(h for _, h in COLS) + " |"
    sep = "|" + "---|" * (len(COLS) + 1)
    lines.append(header)
    lines.append(sep)
    vals = row_values(system_key)
    for model in MODELS:
        lines.append("| " + model + " | " + " | ".join(vals) + " |")
    lines.append("")
    lines.append(f"> {NOTA}")
    lines.append("")

md_path = OUT_DIR / "tablas_profe_nota2.md"
md_path.write_text("\n".join(lines), encoding="utf-8")

# ---------------- Console summary (task step 6) ----------------
order = [("RAG Lexico (BM25)", "BM25 (léxico)"),
         ("RAG Semantico (Dense)", "Denso (BGE)"),
         ("RAG Hibrido Propuesto", "Híbrido (propuesto)")]
print("\n" + "=" * 72)
print("TABLA DE VALORES — retrieval, corpus reconstruido, 194 consultas")
print("=" * 72)
print(f"{'Sistema':<22}{'Prec@1':>9}{'Prec@5':>9}{'Rec@5':>9}{'MRR':>9}{'NDCG@5':>9}")
print("-" * 72)
for key, label in order:
    s = sysd[key]
    print(f"{label:<22}"
          f"{s['precision@1_mean']:>9.3f}"
          f"{s['precision@5_mean']:>9.3f}"
          f"{s['recall@5_mean']:>9.3f}"
          f"{s['mrr_mean']:>9.3f}"
          f"{s['ndcg@5_mean']:>9.3f}")
print("=" * 72)
t = run["timings_seconds"]
print(f"Consultas procesadas : {run['num_queries']}")
print(f"Tiempo retrieval total: {t['total_s']:.1f}s "
      f"(carga modelos {t['index_and_embed_model_load_s'] + t['reranker_load_s']:.1f}s, "
      f"loop {t['retrieval_loop_s']:.1f}s)")
print(f"Tablas escritas:\n  {xlsx_path}\n  {md_path}")
print("=" * 72)
